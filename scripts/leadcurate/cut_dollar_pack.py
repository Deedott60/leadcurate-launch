#!/usr/bin/env python3
"""Cut one paid, single-category Dollar Leads order from a real batch.

The cutter is fail-closed: it refuses sold-out or missing database inventory,
prevents duplicate order-code cuts, and requires a same-day verification
artifact for Fresh Scrub orders.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import sys
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib import error, parse, request


DEFAULT_BATCH_ROOT = Path("/opt/leadcurate/dollar_batches")
DEFAULT_DELIVERY_ROOT = Path("/opt/leadcurate/deliveries/dollar-leads")
PROJECT_URL = "https://jdmlsraqioigbukspduo.supabase.co"
CONTROL_URL = f"{PROJECT_URL}/functions/v1/dollar-fulfillment-control"
ALLOWED_SIZES = {15, 20, 50, 250, 500, 1000}
CODE_RE = re.compile(r"^DL-[A-Z0-9]{4,32}$", re.I)


def clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def load_inventory(batch_root: Path, cycle_slug: str) -> dict[str, Any]:
    path = batch_root / cycle_slug / "inventory.json"
    if not path.exists():
        raise FileNotFoundError(f"inventory manifest not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def lane_manifest(inventory: dict[str, Any], market: str, lane: str) -> dict[str, Any]:
    matches = [item for item in inventory.get("lanes", []) if item.get("market") == market and item.get("lane") == lane]
    if len(matches) != 1:
        raise ValueError(f"expected one inventory lane for {market}/{lane}, found {len(matches)}")
    return matches[0]


def batch_manifest(lane: dict[str, Any], batch_no: int) -> dict[str, Any]:
    matches = [item for item in lane.get("batches", []) if int(item.get("batch_no", 0)) == batch_no]
    if len(matches) != 1:
        raise ValueError(f"batch {batch_no} is not present in real inventory")
    if int(matches[0].get("size", 0)) != 500:
        raise ValueError("batch is not a complete 500-record batch")
    return matches[0]


def read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        fields = list(reader.fieldnames or [])
        rows = list(reader)
    if not fields:
        raise ValueError(f"missing CSV header: {path}")
    return fields, rows


def parcel_key(row: dict[str, str], fields: list[str]) -> str:
    for field in fields:
        value = clean(row.get(field)).upper()
        if value:
            return value
    return ""


def apply_fresh_verification(
    batch_rows: list[dict[str, str]],
    lane: dict[str, Any],
    verified_csv: Path,
    verified_meta: Path,
) -> tuple[list[str], list[dict[str, str]], dict[str, Any]]:
    meta = json.loads(verified_meta.read_text(encoding="utf-8"))
    today = date.today().isoformat()
    if meta.get("verified_at") != today:
        raise ValueError(f"Fresh Scrub verification must be dated {today}")
    if not clean(meta.get("source_url")):
        raise ValueError("Fresh Scrub verification metadata must include source_url")
    if meta.get("market") != lane["market"] or meta.get("lane") != lane["lane"]:
        raise ValueError("Fresh Scrub verification market/lane does not match the order")
    fields, verified_rows = read_rows(verified_csv)
    key_fields = lane["parcel_key_fields"]
    batch_keys = {parcel_key(row, key_fields) for row in batch_rows}
    verified_by_key: dict[str, dict[str, str]] = {}
    for row in verified_rows:
        key = parcel_key(row, key_fields)
        if key and key in batch_keys:
            verified_by_key[key] = row
    ordered = [verified_by_key[parcel_key(row, key_fields)] for row in batch_rows if parcel_key(row, key_fields) in verified_by_key]
    return fields, ordered, meta


def metadata_rows(code: str, lane: dict[str, Any], batch_no: int | str, count: int, fresh: bool, fresh_meta: dict[str, Any] | None) -> list[tuple[str, str]]:
    source_name = clean(fresh_meta.get("source_name")) if fresh_meta else lane["source_name"]
    source_url = clean(fresh_meta.get("source_url")) if fresh_meta else lane["source_url"]
    return [
        ("Dollar Leads Order", code),
        ("Market", lane["market_display"]),
        ("Category", lane["lane_display"]),
        ("County Source", source_name),
        ("Official Source URL", source_url),
        ("Pull Cycle", lane["pull_cycle"]),
        ("Batch Number", str(batch_no)),
        ("Record Count", str(count)),
        ("Fresh Scrub", "yes" if fresh else "no"),
    ]


def write_csv(path: Path, meta_rows: list[tuple[str, str]], fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        for key, value in meta_rows:
            writer.writerow([key, value])
        writer.writerow([])
        writer.writerow(fields)
        for row in rows:
            writer.writerow([row.get(field, "") for field in fields])


def write_xlsx(path: Path, meta_rows: list[tuple[str, str]], fields: list[str], rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write XLSX output") from exc
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = "Dollar Leads"
    for key, value in meta_rows:
        sheet.append([key, value])
    sheet.append([])
    header_row = len(meta_rows) + 2
    sheet.append(fields)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{sheet.cell(row=header_row + len(rows), column=len(fields)).coordinate}"
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(42, max(12, max(len(clean(cell.value)) for cell in column[: min(len(column), 40)]) + 2))
        sheet.column_dimensions[letter].width = width
    workbook.save(path)


class SupabaseInventory:
    def __init__(self, url: str, service_key: str):
        self.url = url.rstrip("/")
        self.service_key = service_key

    def _call(self, method: str, path: str, payload: dict[str, Any] | None = None, prefer: str | None = None) -> Any:
        headers = {"apikey": self.service_key, "Authorization": f"Bearer {self.service_key}", "Content-Type": "application/json"}
        if prefer:
            headers["Prefer"] = prefer
        body = json.dumps(payload).encode("utf-8") if payload is not None else None
        req = request.Request(f"{self.url}/rest/v1/{path}", data=body, method=method, headers=headers)
        try:
            with request.urlopen(req, timeout=30) as response:
                raw = response.read()
                return json.loads(raw) if raw else None
        except error.HTTPError as exc:
            detail = exc.read().decode("utf-8", "replace")
            raise RuntimeError(f"Supabase {method} failed ({exc.code}): {detail}") from exc

    def reserve_seat(self, market: str, lane: str, batch_no: int, cycle: str) -> dict[str, Any]:
        query = parse.urlencode({"market": f"eq.{market}", "lane": f"eq.{lane}", "batch_no": f"eq.{batch_no}", "cycle": f"eq.{cycle}", "status": "eq.live", "select": "id,seats_total,seats_sold"})
        rows = self._call("GET", f"dollar_batches?{query}")
        if not isinstance(rows, list) or len(rows) != 1:
            raise ValueError("live dollar_batches row not found or is ambiguous")
        row = rows[0]
        sold, total = int(row["seats_sold"]), int(row["seats_total"])
        if sold >= total:
            raise ValueError("batch is sold out")
        patch_query = parse.urlencode({"id": f"eq.{row['id']}", "seats_sold": f"eq.{sold}"})
        updated = self._call("PATCH", f"dollar_batches?{patch_query}", {"seats_sold": sold + 1}, "return=representation")
        if not isinstance(updated, list) or len(updated) != 1:
            raise RuntimeError("batch seat changed concurrently; no seat was reserved")
        return updated[0]

    def reserve_founders(self, market: str, lane: str, start_batch_no: int, cycle: str) -> dict[str, Any]:
        result = self._call("POST", "rpc/reserve_dollar_founders_batches", {
            "p_market": market,
            "p_lane": lane,
            "p_start_batch_no": start_batch_no,
            "p_cycle": cycle,
        })
        if not isinstance(result, dict) or len(result.get("batches", [])) != 2:
            raise RuntimeError("Founders reservation did not return two retired batches")
        return result


class FulfillmentControlInventory:
    def __init__(self, url: str, token: str, job_id: str):
        self.url = url
        self.token = token
        self.job_id = job_id

    def reserve(self, batch_no: int) -> dict[str, Any]:
        payload = json.dumps({"action": "reserve", "job_id": self.job_id, "batch_no": batch_no}).encode("utf-8")
        req = request.Request(self.url, data=payload, method="POST", headers={"Content-Type": "application/json", "x-leadcurate-agent-token": self.token})
        try:
            with request.urlopen(req, timeout=45) as response:
                result = json.loads(response.read())
        except error.HTTPError as exc:
            detail = exc.read().decode("utf-8", "replace")
            raise RuntimeError(f"fulfillment reservation failed ({exc.code}): {detail}") from exc
        if not result.get("ok") or not isinstance(result.get("reservation"), dict):
            raise RuntimeError(f"invalid fulfillment reservation response: {result}")
        return result["reservation"]

    def reserve_seat(self, market: str, lane: str, batch_no: int, cycle: str) -> dict[str, Any]:
        return self.reserve(batch_no)

    def reserve_founders(self, market: str, lane: str, start_batch_no: int, cycle: str) -> dict[str, Any]:
        return self.reserve(start_batch_no)


def append_manifest(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, separators=(",", ":")) + "\n")


def main() -> int:
    parser = argparse.ArgumentParser(description="Cut a paid single-category Dollar Leads pack")
    parser.add_argument("--code", required=True)
    parser.add_argument("--market", required=True)
    parser.add_argument("--lane", required=True)
    parser.add_argument("--batch-no", required=True, type=int)
    parser.add_argument("--pack-size", required=True, type=int, choices=sorted(ALLOWED_SIZES))
    parser.add_argument("--cycle-slug", default=datetime.now().strftime("%Y-%m"))
    parser.add_argument("--batch-root", type=Path, default=DEFAULT_BATCH_ROOT)
    parser.add_argument("--delivery-root", type=Path, default=DEFAULT_DELIVERY_ROOT)
    parser.add_argument("--fresh", action="store_true")
    parser.add_argument("--fresh-verified-csv", type=Path)
    parser.add_argument("--fresh-verified-meta", type=Path)
    parser.add_argument("--dry-run", action="store_true", help="Write to the chosen delivery root without changing dollar_batches")
    parser.add_argument("--job-id", help="Fulfillment job UUID for the guarded VPS reservation path")
    args = parser.parse_args()

    code = args.code.upper()
    if not CODE_RE.fullmatch(code):
        raise ValueError("order code must match DL- followed by 4 to 32 letters or digits")
    if args.fresh and args.pack_size != 20:
        raise ValueError("Fresh Scrub orders must contain exactly 20 records")
    if not args.fresh and args.pack_size == 20:
        raise ValueError("20-record packs require --fresh")
    if args.fresh and (not args.fresh_verified_csv or not args.fresh_verified_meta):
        raise ValueError("Fresh Scrub requires --fresh-verified-csv and --fresh-verified-meta")

    inventory = load_inventory(args.batch_root, args.cycle_slug)
    lane = lane_manifest(inventory, args.market, args.lane)
    founder = args.pack_size == 1000
    batches = [batch_manifest(lane, args.batch_no)]
    if founder:
        batches.append(batch_manifest(lane, args.batch_no + 1))
    fields, batch_rows = read_rows(Path(batches[0]["file"]))
    if len(batch_rows) != 500:
        raise ValueError("batch file does not contain exactly 500 records")
    if founder:
        second_fields, second_rows = read_rows(Path(batches[1]["file"]))
        if second_fields != fields or len(second_rows) != 500:
            raise ValueError("second Founders batch is incomplete or has a different schema")
        keys = lane["parcel_key_fields"]
        first_keys = {parcel_key(row, keys) for row in batch_rows}
        if first_keys.intersection(parcel_key(row, keys) for row in second_rows):
            raise ValueError("Founders batches overlap")
        batch_rows.extend(second_rows)
    fresh_meta = None
    if args.fresh:
        fields, batch_rows, fresh_meta = apply_fresh_verification(batch_rows, lane, args.fresh_verified_csv, args.fresh_verified_meta)
    if len(batch_rows) < args.pack_size:
        raise ValueError(f"only {len(batch_rows)} eligible records remain after verification")
    selected = batch_rows[: args.pack_size]

    final_dir = args.delivery_root / code
    if final_dir.exists():
        raise FileExistsError(f"delivery already exists for {code}: {final_dir}")
    args.delivery_root.mkdir(parents=True, exist_ok=True)
    pending = Path(tempfile.mkdtemp(prefix=f".{code}-pending-", dir=args.delivery_root))
    try:
        batch_label = f"{args.batch_no}-{args.batch_no + 1}" if founder else args.batch_no
        meta_rows = metadata_rows(code, lane, batch_label, len(selected), args.fresh, fresh_meta)
        csv_path = pending / f"{code}-{args.market}-{args.lane}-{args.pack_size}.csv"
        xlsx_path = pending / f"{code}-{args.market}-{args.lane}-{args.pack_size}.xlsx"
        write_csv(csv_path, meta_rows, fields, selected)
        write_xlsx(xlsx_path, meta_rows, fields, selected)
        seat = None
        if not args.dry_run:
            service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
            if service_key:
                inventory_api: Any = SupabaseInventory(os.environ.get("SUPABASE_URL", PROJECT_URL), service_key)
            else:
                automation_token = os.environ.get("HOSTINGER_WEBHOOK_SECRET", "")
                if not automation_token or not args.job_id:
                    raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY or HOSTINGER_WEBHOOK_SECRET plus --job-id is required unless --dry-run is used")
                inventory_api = FulfillmentControlInventory(os.environ.get("DOLLAR_FULFILLMENT_CONTROL_URL", CONTROL_URL), automation_token, args.job_id)
            seat = inventory_api.reserve_founders(args.market, args.lane, args.batch_no, lane["pull_cycle"]) if founder else inventory_api.reserve_seat(args.market, args.lane, args.batch_no, lane["pull_cycle"])
        manifest = {
            "order_code": code, "market": args.market, "market_display": lane["market_display"],
            "lane": args.lane, "lane_display": lane["lane_display"], "batch_no": args.batch_no,
            "batch_nos": [args.batch_no, args.batch_no + 1] if founder else [args.batch_no],
            "pack_size": args.pack_size, "record_count": len(selected), "cycle": lane["pull_cycle"],
            "source_name": fresh_meta.get("source_name", lane["source_name"]) if fresh_meta else lane["source_name"],
            "source_url": fresh_meta.get("source_url", lane["source_url"]) if fresh_meta else lane["source_url"],
            "fresh_scrub": args.fresh, "fresh_verification": fresh_meta,
            "batch_file": batches[0]["file"], "batch_sha256": batches[0]["sha256"],
            "batch_files": [batch["file"] for batch in batches],
            "batch_sha256s": [batch["sha256"] for batch in batches],
            "csv": csv_path.name, "xlsx": xlsx_path.name,
            "seat_after_cut": int(seat["seats_sold"]) if seat and not founder else None,
            "founders_slots_sold": int(seat["promo_slots_sold"]) if seat and founder else None,
            "dry_run": args.dry_run, "cut_at_utc": datetime.now(timezone.utc).isoformat(),
        }
        (pending / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        pending.rename(final_dir)
        manifest["delivery_dir"] = str(final_dir)
        append_manifest(args.delivery_root / "manifest.jsonl", manifest)
        print(json.dumps(manifest, indent=2))
    except Exception:
        shutil.rmtree(pending, ignore_errors=True)
        raise
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
