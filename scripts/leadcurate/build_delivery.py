#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import re
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Callable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAW_ROOT = Path("/opt/leadcurate/raw_imports")
SNAPSHOT_ROOT = Path("/opt/leadcurate/snapshots")

LANE_LABELS = {
    "tax-delinquent": "Tax Delinquent",
    "probate": "Probate Premium",
    "pre-foreclosure": "Pre-Foreclosure Premium",
    "code-violations": "Code Violations List",
    "liens": "Liens Watchlist",
    "absentee": "Absentee Owner List",
    "active-permits": "Active Permits Distress",
    "high-equity": "High-Equity Owners List",
    "individual-homeowner": "Active Homeowner List",
    "entity-owned": "Entity-Owned Properties",
    "vacant-land": "Vacant Land",
}

MARKET_REGISTRY: dict[str, dict[str, Any]] = {
    "wake-nc": {"display": "Wake County NC", "raw_dir": RAW_ROOT / "wake-nc", "raw_pattern": "delinquent*.xlsx", "lane_patterns": {"default": ["parcels.csv", "property.csv"], "tax-delinquent": ["delinquent*.xlsx"], "absentee": ["parcels.csv", "property.csv"], "high-equity": ["parcels.csv", "property.csv"], "individual-homeowner": ["parcels.csv", "property.csv"], "entity-owned": ["parcels.csv", "property.csv"], "vacant-land": ["parcels.csv", "property.csv"]}, "default_city": "Raleigh", "state": "NC"},
    "cobb-ga": {"display": "Cobb County GA", "raw_dir": RAW_ROOT / "cobb-ga", "raw_pattern": "*.pdf", "lane_patterns": {"tax-delinquent": ["*.pdf"], "default": ["*.pdf"]}, "default_city": "Marietta", "state": "GA", "snapshot_pattern": "cobb-ga-delinquent-*.csv"},
    "guilford-nc": {"display": "Guilford County NC", "raw_dir": RAW_ROOT / "guilford-nc", "raw_pattern": "tax-delinquent-report.csv", "lane_patterns": {"tax-delinquent": ["tax-delinquent-report.csv"], "pre-foreclosure": ["parcel-foreclosure.csv"], "default": ["county-parcels.csv", "historical-parcels-2025.csv"]}, "default_city": "Greensboro", "state": "NC"},
    "marion-in": {"display": "Marion County IN", "raw_dir": RAW_ROOT / "marion-in", "raw_pattern": "parcels-owner-assessed.csv", "lane_patterns": {"default": ["parcels-owner-assessed.csv", "parcels-base.csv", "hhc-parcel-owner.csv"]}, "default_city": "Indianapolis", "state": "IN", "snapshot_pattern": "marion-in-*.csv"},
    "dekalb-ga": {"display": "DeKalb County GA", "raw_dir": RAW_ROOT / "dekalb-ga", "raw_pattern": "tax-parcels-2025.csv", "lane_patterns": {"default": ["tax-parcels-2025.csv"]}, "default_city": "Decatur", "state": "GA", "snapshot_pattern": "dekalb-ga-*.csv"},
    "forsyth-nc": {"display": "Forsyth County NC", "raw_dir": RAW_ROOT / "forsyth-nc", "raw_pattern": "parcels.csv", "lane_patterns": {"default": ["parcels.csv", "parcels-hosted.csv"], "pre-foreclosure": ["bank-foreclosures.csv"]}, "default_city": "Winston-Salem", "state": "NC", "snapshot_pattern": "forsyth-nc-*.csv"},
    "fulton-ga": {"display": "Fulton County GA", "raw_dir": RAW_ROOT / "fulton-ga", "raw_pattern": "tax-parcels-2025.csv", "lane_patterns": {"default": ["tax-parcels-2025.csv", "current-parcels.csv", "parcels.csv"]}, "default_city": "Atlanta", "state": "GA", "snapshot_pattern": "fulton-ga-*.csv"},
    "harris-tx": {"display": "Harris County TX", "raw_dir": RAW_ROOT / "harris-tx", "raw_pattern": "real_acct.txt", "lane_patterns": {"default": ["real_acct.txt"], "active-permits": ["permits.txt", "real_acct.txt"]}, "default_city": "Houston", "state": "TX", "snapshot_pattern": "harris-tx-permit-burnout-*.csv"},
    "jefferson-al": {"display": "Jefferson County AL", "raw_dir": RAW_ROOT / "jefferson-al", "raw_pattern": "DelinquentParcelList.xls", "lane_patterns": {"tax-delinquent": ["DelinquentParcelList.xls"], "default": ["DelinquentParcelList.xls"]}, "default_city": "Birmingham", "state": "AL", "snapshot_pattern": "jefferson-al-delinquent-*.csv"},
    "mecklenburg-nc": {"display": "Mecklenburg County NC", "raw_dir": RAW_ROOT / "mecklenburg-nc", "raw_pattern": "parcel-lookup.csv", "lane_patterns": {"default": ["parcel-lookup.csv", "parcels-full.csv"], "probate": ["mecklenburg-probate.csv"], "liens": ["lien-data.csv"], "vacant-land": ["vacant-land.csv"], "absentee": ["parcel-lookup.csv", "parcels-full.csv"], "high-equity": ["parcel-lookup.csv", "parcels-full.csv"], "individual-homeowner": ["parcel-lookup.csv", "parcels-full.csv"], "entity-owned": ["parcel-lookup.csv", "parcels-full.csv"]}, "default_city": "Charlotte", "state": "NC"},
    "cuyahoga-oh": {"display": "Cuyahoga County OH", "raw_dir": RAW_ROOT / "cuyahoga-oh", "raw_pattern": "tax-parcels.csv", "lane_patterns": {"default": ["tax-parcels.csv"], "high-equity": ["tax-parcels.csv"], "absentee": ["tax-parcels.csv"], "individual-homeowner": ["tax-parcels.csv"], "entity-owned": ["tax-parcels.csv"], "vacant-land": ["tax-parcels.csv"]}, "default_city": "Cleveland", "state": "OH"},
    "tarrant-tx": {"display": "Tarrant County TX", "raw_dir": RAW_ROOT / "tarrant-tx", "raw_pattern": "tax-roll.zip", "lane_patterns": {"default": ["tax-roll.zip"], "tax-delinquent": ["tax-roll.zip"]}, "default_city": "Fort Worth", "state": "TX"},
    "jefferson-ky": {"display": "Jefferson County KY", "raw_dir": RAW_ROOT / "jefferson-ky", "raw_pattern": "parcels.csv", "lane_patterns": {"default": ["parcels.csv"], "code-violations": ["property-maintenance-violations.csv"], "liens": ["lien-holder-final-orders.csv"], "pre-foreclosure": ["property-foreclosures.csv"]}, "default_city": "Louisville", "state": "KY"},
    "shelby-tn": {"display": "Shelby County TN", "raw_dir": RAW_ROOT / "shelby-tn", "raw_pattern": "tax-sale-extract.csv", "lane_patterns": {"default": ["tax-sale-extract.csv"], "tax-delinquent": ["tax-sale-extract.csv"], "pre-foreclosure": ["tax-sale-extract.csv"]}, "default_city": "Memphis", "state": "TN"},
    "duval-fl": {"display": "Duval County FL", "raw_dir": RAW_ROOT / "duval-fl", "raw_pattern": "parcels.csv", "lane_patterns": {"default": ["parcels.csv"], "tax-delinquent": ["tax-delinquent*.csv", "parcels.csv"], "pre-foreclosure": ["pre-foreclosure*.csv", "foreclosure*.csv"], "code-violations": ["code-violations*.csv"], "liens": ["liens*.csv"], "active-permits": ["permits*.csv"], "absentee": ["parcels.csv"], "high-equity": ["parcels.csv"], "individual-homeowner": ["parcels.csv"], "entity-owned": ["parcels.csv"], "vacant-land": ["parcels.csv"]}, "default_city": "Jacksonville", "state": "FL"},
    "davidson-tn": {"display": "Davidson County TN", "raw_dir": RAW_ROOT / "davidson-tn", "raw_pattern": "parcels.csv", "lane_patterns": {"default": ["parcels.csv"], "tax-delinquent": ["tax-delinquent*.csv", "parcels.csv"], "pre-foreclosure": ["pre-foreclosure*.csv", "foreclosure*.csv"], "code-violations": ["code-violations*.csv"], "liens": ["liens*.csv"], "active-permits": ["permits*.csv"], "absentee": ["parcels.csv"], "high-equity": ["parcels.csv"], "individual-homeowner": ["parcels.csv"], "entity-owned": ["parcels.csv"], "vacant-land": ["parcels.csv"]}, "default_city": "Nashville", "state": "TN"},
    "maricopa-az": {"display": "Maricopa County AZ", "raw_dir": RAW_ROOT / "maricopa-az", "raw_pattern": "residential-master.zip", "lane_patterns": {"default": ["residential-master.zip", "secured-master.zip"], "individual-homeowner": ["residential-master.zip"], "high-equity": ["residential-master.zip", "secured-master.zip"], "entity-owned": ["commercial-master.zip", "secured-master.zip"], "vacant-land": ["secured-master.zip"]}, "default_city": "Phoenix", "state": "AZ"},
    "allen-in": {"display": "Allen County IN", "raw_dir": RAW_ROOT / "allen-in", "raw_pattern": "2025-delinquent-property.xlsx", "lane_patterns": {"default": ["2025-delinquent-property.xlsx"], "tax-delinquent": ["2025-delinquent-property.xlsx"]}, "default_city": "Fort Wayne", "state": "IN"},
    "charleston-sc": {"display": "Charleston County SC", "raw_dir": RAW_ROOT / "charleston-sc", "raw_pattern": "*Tax-Sale-Listing.pdf", "lane_patterns": {"default": ["*Tax-Sale-Listing.pdf"], "tax-delinquent": ["*Tax-Sale-Listing.pdf"]}, "default_city": "Charleston", "state": "SC"},
    "greenville-sc": {"display": "Greenville County SC", "raw_dir": RAW_ROOT / "greenville-sc", "raw_pattern": "tax-sale-info.pdf", "lane_patterns": {"default": ["tax-sale-info.pdf", "tax-sale-app.html"], "tax-delinquent": ["tax-sale-info.pdf", "tax-sale-app.html"]}, "default_city": "Greenville", "state": "SC"},
    "dallas-tx": {"display": "Dallas County TX", "raw_dir": RAW_ROOT / "dallas-tx", "raw_pattern": "2025-real-property-cert-roll.zip", "lane_patterns": {"default": ["2025-real-property-cert-roll.zip", "parcel2025.zip"], "tax-delinquent": ["2025-real-property-cert-roll.zip"], "vacant-land": ["parcel2025.zip"]}, "default_city": "Dallas", "state": "TX"},
    "erie-ny": {"display": "Erie County NY", "raw_dir": RAW_ROOT / "erie-ny", "raw_pattern": "*.pdf", "lane_patterns": {"default": ["*.pdf"], "tax-delinquent": ["*delinquent*.pdf", "*.pdf"], "pre-foreclosure": ["*foreclosure*.pdf", "*.pdf"]}, "default_city": "Buffalo", "state": "NY"},
    "fayette-ky": {"display": "Fayette County KY", "raw_dir": RAW_ROOT / "fayette-ky", "raw_pattern": "parcel.csv", "lane_patterns": {"default": ["parcel.csv"], "vacant-land": ["vacant-land-2010.csv"], "individual-homeowner": ["parcel.csv"]}, "default_city": "Lexington", "state": "KY"},
    "nyc": {"display": "New York City NY", "raw_dir": RAW_ROOT / "nyc", "raw_pattern": "tax-lien.csv", "lane_patterns": {"default": ["tax-lien.csv"], "tax-delinquent": ["tax-lien.csv", "hpd-tax-delinquency.csv"], "liens": ["tax-lien.csv"], "code-violations": ["dob-violations.csv"]}, "default_city": "New York", "state": "NY"},
}

ENTITY_WORDS = (" LLC", " INC", " CORP", " LP", " TTC", " FUND", " TRUST", " L P", " L L C", " LLP", " COMPANY", " PROPERTIES", " INVESTMENTS", " HOLDINGS", " PARTNERS", " CHURCH", " CITY OF", " COUNTY")


def clean(value: Any) -> str:
    return str(value or "").strip()


def money(value: Any) -> float:
    text = clean(value).replace("$", "").replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def is_residential_owner(owner: str, allow_entities: bool = False) -> bool:
    if not owner:
        return False
    if allow_entities:
        return True
    upper = f" {owner.upper()} "
    return not any(word in upper for word in ENTITY_WORDS)


def parse_mailing_csz(addr2: str) -> tuple[str, str, str]:
    text = clean(addr2)
    if not text:
        return "", "", ""
    zip_match = re.search(r"\b(\d{5}(?:-\d{4})?)\s*$", text)
    zip_code = zip_match.group(1) if zip_match else ""
    rest = text[: zip_match.start()].strip().rstrip(",") if zip_match else text
    state_match = re.search(r"\b([A-Z]{2})\s*$", rest)
    state = state_match.group(1) if state_match else ""
    city = rest[: state_match.start()].strip().rstrip(",") if state_match else rest
    return city.title(), state, zip_code


def motivation(owed: float, years: int) -> str:
    if owed >= 10000 or years >= 3:
        return "HOT"
    if owed >= 3000 or years >= 2:
        return "WARM"
    return "WORKING"


def date_dirs(raw_dir: Path) -> list[Path]:
    if not raw_dir.exists():
        return []
    dirs = [p for p in raw_dir.iterdir() if p.is_dir()]
    return sorted(dirs, key=lambda p: p.name, reverse=True)


def lane_patterns(cfg: dict[str, Any], lane: str) -> list[str]:
    raw = cfg.get("lane_patterns") or {}
    patterns = raw.get(lane) or raw.get("default") or [cfg.get("raw_pattern", "*")]
    if isinstance(patterns, str):
        patterns = [patterns]
    fallback = cfg.get("raw_pattern")
    if fallback and fallback not in patterns:
        patterns.append(fallback)
    return patterns


def latest_file(market: str, lane: str) -> Path:
    cfg = MARKET_REGISTRY[market]
    raw_dir = Path(cfg["raw_dir"])
    patterns = lane_patterns(cfg, lane)
    for dated in date_dirs(raw_dir):
        for pattern in patterns:
            matches = sorted(dated.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
            if matches:
                return matches[0]
    for pattern in patterns:
        matches = sorted(raw_dir.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
        if matches:
            return matches[0]
    raise FileNotFoundError(f"No raw file found for {market}/{lane}")


def latest_snapshot(market: str, pattern: str | None = None) -> Path | None:
    root = SNAPSHOT_ROOT / market
    if not root.exists():
        return None
    patterns = [pattern or "*.csv"]
    candidates: list[Path] = []
    for dated in date_dirs(root):
        for pat in patterns:
            candidates.extend(p for p in dated.glob(pat) if "preview" not in p.name and "tiered" not in p.name)
    return sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0] if candidates else None


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        return list(csv.DictReader(handle))


def read_xlsx_rows(path: Path) -> tuple[list[str], list[tuple[Any, ...]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(h) for h in next(rows)]
    return headers, list(rows)


def read_xlsx_dicts(path: Path) -> list[dict[str, Any]]:
    headers, rows = read_xlsx_rows(path)
    return [dict(zip(headers, row)) for row in rows]


def first_present(raw: dict[str, Any], *names: str) -> Any:
    for name in names:
        value = raw.get(name)
        if clean(value):
            return value
    return ""


def join_present(*values: Any) -> str:
    return " ".join(clean(v) for v in values if clean(v))


def is_entity_owner(owner: str) -> bool:
    upper = f" {owner.upper()} "
    return any(word in upper for word in ENTITY_WORDS)


def normalize_record(raw: dict[str, Any], cfg: dict[str, Any], lane_label: str, lane: str = "tax-delinquent") -> dict[str, Any] | None:
    owner = clean(
        first_present(
            raw,
            "owner", "owner_name", "Owner Name", "OWNER_NAME", "OWNERNME1",
            "FULLOWNERNAME", "CURRENTOWNERNAME1", "parcel_owner", "mail_name",
            "Pay Yr Owner Of Record", "Customer_Name", "fullname",
        )
        or join_present(raw.get("Owner_FirstName"), raw.get("Owner_LastName"))
        or join_present(raw.get("ownerfirstname"), raw.get("ownerlastname"))
    )
    allow_entities = lane == "entity-owned"
    if not is_residential_owner(owner, allow_entities=allow_entities):
        return None
    parcel = clean(first_present(raw, "parcel", "parcel_id", "Parcel REID", "PARCEL_NUM", "PARCELID", "ParcelID", "PID", "Tax_ID", "taxpid", "pid", "nc_pin", "Parcel/Property Number", "PARCEL_ID", "Full_Parcel_ID", "PVANUM", "Block ", "BLOCK"))
    account = clean(first_present(raw, "account", "Account ID", "acct", "ACCOUNT_NUM", "LienNo", "InvoiceNo", "Case_", "b1_alt_id") or parcel)
    if not (parcel or account):
        return None
    total_owed = money(first_present(raw, "total_owed", "Total Owed", "TOTAL_DUE_AMOUNT", "delinquent_amount", "total_due", "BILL_DUE_AMT", "score", "Delinquent Amt", "Delinquent Amt ", "CitationAmount", "final_citation_amount", "Sale_Price", "Water Debt Only"))
    value = money(first_present(raw, "value", "Property Value", "PROP_ASSESS_VALUE", "TOT_APPR", "APPRAISED_VALUE", "CAMAPARCELID", "mkt_val", "Total_Value", "TOTALVALUE", "tax_market_total", "certified_tax_total", "totalvalue", "Total Value"))
    source_signal_lanes = {"pre-foreclosure", "code-violations", "liens", "probate"}
    if value <= 0 and total_owed <= 0 and lane not in source_signal_lanes:
        return None
    years = int(money(first_present(raw, "years", "Years Behind", "tax_years", "Cycle") or 1)) or 1
    prop_zip = clean(first_present(raw, "Property ZIP", "SITEZIP", "ZIPCODE", "site_zip", "Zip_Code", "parcel_zip", "Zip", "zip", "Zip Code"))
    mail_zip = clean(first_present(raw, "Mailing ZIP", "MAIL_ZIP", "PSTLZIP5", "OWNERZIP", "mail_zip", "Zip_Code", "mail_zip", "zipcode"))
    absentee = "Yes" if (mail_zip and prop_zip and mail_zip[:5] != prop_zip[:5]) or clean(raw.get("absentee")).upper() in ("Y", "YES", "TRUE") else "No"
    address = clean(
        first_present(raw, "address", "Property Address", "SITEADDRESS", "site_addr", "PROPERTYADDRESS", "Location", "parcel_addr", "FullAddress", "PartialAddress", "Property_Address", "ADDRESS", "FULL_ADDRESS")
        or join_present(raw.get("House_Nr"), raw.get("Dir"), raw.get("Street_Name"), raw.get("St_Type"), raw.get("Post_Dir"))
        or join_present(raw.get("Street Number"), raw.get("Street Name"))
        or join_present(raw.get("House Number"), raw.get("Street Name"))
    )
    city = clean(first_present(raw, "city", "Property City", "SITECITY", "site_city", "CITY", "parcel_city", "municipality") or cfg.get("default_city"))
    equity = max(0.0, value - total_owed)
    rec = {
        "Account ID": account,
        "Parcel REID": parcel,
        "Owner Name": owner.title(),
        "Secondary Owner": clean(first_present(raw, "secondary", "Secondary Owner", "OWNERNME2", "second_owner", "cownerfirstname")),
        "Property Address": address.title(),
        "Property City": city.title(),
        "Property ZIP": prop_zip,
        "Mailing Street": clean(first_present(raw, "Mailing Street", "MAIL_ADDR1", "PSTLADDRESS", "OWNERADDRESS", "mail_address_1", "mail_addr_street", "Mailing_Address", "mailaddr1", "addr1")).title(),
        "Mailing City": clean(first_present(raw, "Mailing City", "MAIL_CITY", "PSTLCITY", "OWNERCITY", "mail_city", "city")).title(),
        "Mailing State": clean(first_present(raw, "Mailing State", "MAIL_STATE", "PSTLSTATE", "OWNERSTATE", "mail_state", "state")),
        "Mailing ZIP": mail_zip,
        "Absentee Owner": absentee,
        "Acres": round(money(first_present(raw, "Acres", "PROP_SIZE", "ACREAGE", "Total_Acreage", "parcel_acreage", "totalac", "PVA_ACRE")), 2),
        "Property Value": round(value, 2),
        "Building Value": round(money(first_present(raw, "Building Value", "IMPR_APPR", "Building_Value", "bld_val", "Building_Value", "certified_tax_building", "netbldgvalue")), 2),
        "Land Value": round(money(first_present(raw, "Land Value", "LNDVALUE", "Land_Value", "land_val", "certified_tax_land", "landvalue")), 2),
        "Years Behind": years,
        "Total Owed": round(total_owed, 2),
        "Estimated Equity": round(equity, 2),
        "Motivation": motivation(total_owed, years),
        "Lane": lane_label,
    }
    if not lane_match(raw, rec, lane):
        return None
    return rec


def lane_match(raw: dict[str, Any], rec: dict[str, Any], lane: str) -> bool:
    text = " ".join(clean(v) for v in raw.values()).lower()
    if lane in {"tax-delinquent", "active-permits", "probate", "pre-foreclosure", "code-violations", "liens"}:
        return True
    if lane == "absentee":
        return rec["Absentee Owner"] == "Yes"
    if lane == "high-equity":
        return float(rec["Estimated Equity"]) >= 100000
    if lane == "individual-homeowner":
        return not is_entity_owner(rec["Owner Name"])
    if lane == "entity-owned":
        return is_entity_owner(rec["Owner Name"])
    if lane == "vacant-land":
        return (
            float(rec["Building Value"]) <= 0
            or "vacant" in text
            or "land" in text
            or clean(raw.get("vacantorimproved")).lower() == "vacant"
        )
    return True


def parse_wake_nc(path: Path, cfg: dict[str, Any], lane: str) -> list[dict[str, Any]]:
    headers, rows = read_xlsx_rows(path)
    h = {name: i for i, name in enumerate(headers)}
    parcels: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        td = money(row[h.get("TOTAL_DUE")])
        tv = money(row[h.get("TOTAL_REAL_VALUE")])
        owner = clean(row[h.get("Primary_Owner")])
        if td <= 0 or tv < 10000 or not is_residential_owner(owner):
            continue
        street_name = clean(row[h.get("STREET_NAME")])
        reid = clean(row[h.get("REID")])
        acct = clean(row[h.get("ACCOUNT_NUM")])
        key = (reid, acct)
        tax_year = int(money(row[h.get("TAX_YEAR")]) or 2025)
        if key not in parcels:
            parcels[key] = {"tax_years": set(), "total_owed": 0.0}
        p = parcels[key]
        p.update({
            "parcel": reid,
            "account": acct,
            "owner": owner,
            "secondary": clean(row[h.get("Secondary_Owner")]).title(),
            "address": f"{clean(row[h.get('Street_Number')])} {street_name} {clean(row[h.get('ST_TYPE')])}".strip(),
            "city": clean(row[h.get("CITY")]) or cfg["default_city"],
            "Property ZIP": clean(row[h.get("ZIP_CODE")]).split(".")[0],
            "Acres": money(row[h.get("ACRES")]),
            "value": tv,
            "Building Value": money(row[h.get("Building_value")]),
            "Land Value": money(row[h.get("LAND_VALUE")]),
            "Mailing Street": clean(row[h.get("Mailing_Address1")]),
        })
        c, s, z = parse_mailing_csz(clean(row[h.get("Mailing_Address2")]))
        p["Mailing City"], p["Mailing State"], p["Mailing ZIP"] = c, s, z
        p["tax_years"].add(tax_year)
        p["total_owed"] += td
    out = []
    for p in parcels.values():
        p["years"] = len(p.pop("tax_years"))
        rec = normalize_record(p, cfg, LANE_LABELS.get(lane, "Tax Delinquent"), lane)
        if rec:
            out.append(rec)
    return out


def parse_guilford_nc(path: Path, cfg: dict[str, Any], lane: str) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in read_csv(path):
        parcel = clean(row.get("PARCEL_NUM"))
        if not parcel:
            continue
        rec = grouped.setdefault(parcel, dict(row, tax_years=set(), total_owed=0.0))
        rec["total_owed"] += money(row.get("TOTAL_DUE_AMOUNT"))
        rec["tax_years"].add(clean(row.get("TAX_YEAR")))
    out = []
    for rec in grouped.values():
        rec["years"] = len(rec.pop("tax_years"))
        normalized = normalize_record(rec, cfg, LANE_LABELS.get(lane, "Tax Delinquent"), lane)
        if normalized:
            out.append(normalized)
    return out


def parse_snapshot(path: Path, cfg: dict[str, Any], lane: str) -> list[dict[str, Any]]:
    label = LANE_LABELS.get(lane, "Tax Delinquent")
    return [r for r in (normalize_record(row, cfg, label, lane) for row in read_csv(path)) if r]


def parse_generic_csv(path: Path, cfg: dict[str, Any], lane: str) -> list[dict[str, Any]]:
    label = LANE_LABELS.get(lane, "Tax Delinquent")
    return [r for r in (normalize_record(row, cfg, label, lane) for row in read_csv(path)) if r]


def parse_generic_xlsx(path: Path, cfg: dict[str, Any], lane: str) -> list[dict[str, Any]]:
    label = LANE_LABELS.get(lane, "Tax Delinquent")
    return [r for r in (normalize_record(row, cfg, label, lane) for row in read_xlsx_dicts(path)) if r]


PARSERS: dict[str, Callable[[Path, dict[str, Any], str], list[dict[str, Any]]]] = {
    "wake-nc": parse_wake_nc,
    "guilford-nc": parse_guilford_nc,
}


def load_records(market: str, lane: str) -> tuple[Path, list[dict[str, Any]]]:
    cfg = MARKET_REGISTRY[market]
    if market in PARSERS and lane == "tax-delinquent":
        src = latest_file(market, lane)
        return src, PARSERS[market](src, cfg, lane)
    snap = latest_snapshot(market, cfg.get("snapshot_pattern"))
    if snap:
        return snap, parse_snapshot(snap, cfg, lane)
    src = latest_file(market, lane)
    suffix = src.suffix.lower()
    if suffix == ".csv" or suffix in {".txt", ".dat"}:
        return src, parse_generic_csv(src, cfg, lane)
    if suffix in {".xlsx", ".xlsm"}:
        return src, parse_generic_xlsx(src, cfg, lane)
    raise SystemExit(
        f"Unsupported raw source for {market}/{lane}: {src}. "
        "Workflow should route this to LLM/manual review or an extractor module before delivery."
    )


def add_analytics(records: list[dict[str, Any]], market: str, lane: str, src: Path) -> dict[str, Any]:
    hot = sum(1 for r in records if r["Motivation"] == "HOT")
    warm = sum(1 for r in records if r["Motivation"] == "WARM")
    absentee = sum(1 for r in records if r["Absentee Owner"] == "Yes")
    heirs = sum(1 for r in records if re.search(r"\b(heirs|hrs)\b", r["Owner Name"], re.I))
    owed = [float(r["Total Owed"]) for r in records]
    years = [int(r["Years Behind"]) for r in records]
    return {
        "market_slug": market,
        "market": MARKET_REGISTRY[market]["display"],
        "lane": lane,
        "source_file": str(src),
        "total": len(records),
        "hot": hot,
        "warm": warm,
        "working": len(records) - hot - warm,
        "absentee": absentee,
        "heirs_count": heirs,
        "top_equity": max((float(r["Estimated Equity"]) for r in records), default=0),
        "top_debt": max(owed, default=0),
        "median_debt": sorted(owed)[len(owed) // 2] if owed else 0,
        "median_years": sorted(years)[len(years) // 2] if years else 0,
        "avg_debt": round(sum(owed) / len(owed), 2) if owed else 0,
        "debt_buckets": bucketize(owed, [1000, 3000, 5000, 10000], ["<$1K", "$1K-$3K", "$3K-$5K", "$5K-$10K", "$10K+"]),
        "years_buckets": bucketize(years, [1, 2, 3, 5], ["0-1", "2", "3", "4-5", "6+"]),
    }


def bucketize(values: list[float], thresholds: list[float], labels: list[str]) -> list[dict[str, Any]]:
    buckets = [{"label": label, "value": 0} for label in labels]
    for value in values:
        idx = 0
        while idx < len(thresholds) and value > thresholds[idx]:
            idx += 1
        buckets[idx]["value"] += 1
    return buckets


def write_outputs(records: list[dict[str, Any]], summary: dict[str, Any], output_dir: Path, count: int) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    slug = summary["market_slug"]
    xlsx_path = output_dir / f"{slug}-Curated-Distress-{count}.xlsx"
    json_path = output_dir / f"{slug}-preview.json"
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "LeadCurate"
    cover["A1"].font = Font(size=24, bold=True, color="15803d")
    cover["A3"] = f"Curated Distress List - {summary['market']}"
    cover["A3"].font = Font(size=16, bold=True)
    for i, (key, value) in enumerate([
        ("Market:", summary["market"]),
        ("Lane:", summary["lane"]),
        ("Records:", summary["total"]),
        ("HOT:", summary["hot"]),
        ("WARM:", summary["warm"]),
        ("Absentee owners:", summary["absentee"]),
        ("Top equity:", f"${summary['top_equity']:,.2f}"),
        ("Source:", summary["source_file"]),
    ], 5):
        cover[f"A{i}"] = key
        cover[f"B{i}"] = value
    sheet = wb.create_sheet("Records")
    cols = list(records[0].keys()) if records else []
    sheet.append(cols)
    for rec in records:
        sheet.append([rec.get(c) for c in cols])
    fill = PatternFill("solid", fgColor="15803d")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    for idx, col in enumerate(cols, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = max(12, min(32, len(col) + 2))
    sheet.freeze_panes = "A2"
    wb.save(xlsx_path)
    payload = dict(summary)
    payload["preview"] = [
        {"owner": r["Owner Name"], "address": r["Property Address"], "city": r["Property City"], "owed": r["Total Owed"], "value": r["Property Value"], "equity": r["Estimated Equity"], "years": r["Years Behind"], "motivation": r["Motivation"], "absentee": r["Absentee Owner"]}
        for r in records[:30]
    ]
    json_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return {"xlsx": str(xlsx_path), "json": str(json_path)}


def dedupe_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    best: dict[str, dict[str, Any]] = {}
    for rec in records:
        key = f"{rec.get('Parcel REID', '')}|{rec.get('Account ID', '')}"
        if key not in best:
            best[key] = rec
            continue
        current = best[key]
        if (float(rec.get("Total Owed") or 0), float(rec.get("Estimated Equity") or 0)) > (
            float(current.get("Total Owed") or 0),
            float(current.get("Estimated Equity") or 0),
        ):
            best[key] = rec
    return list(best.values())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--market", required=True, choices=sorted(MARKET_REGISTRY))
    parser.add_argument("--lane", default="tax-delinquent")
    parser.add_argument("--count", type=int, default=500)
    parser.add_argument("--output-dir", default="/tmp")
    args = parser.parse_args()
    src, records = load_records(args.market, args.lane)
    records = dedupe_records(records)
    records.sort(key=lambda r: (-float(r["Total Owed"]), -float(r["Estimated Equity"])))
    top = records[: args.count]
    if not top:
        raise SystemExit(f"No deliverable records built for {args.market}/{args.lane} from {src}")
    summary = add_analytics(top, args.market, args.lane, src)
    outputs = write_outputs(top, summary, Path(args.output_dir), args.count)
    print(json.dumps({"ok": True, **summary, **outputs}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
