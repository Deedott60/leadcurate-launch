#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path
from typing import Any

import openpyxl

import build_delivery

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_ROOT = Path("/opt/leadcurate/deliveries")


def run(cmd: list[str]) -> dict[str, Any]:
    proc = subprocess.run(cmd, text=True, capture_output=True)
    return {
        "cmd": cmd,
        "returncode": proc.returncode,
        "stdout": proc.stdout.strip(),
        "stderr": proc.stderr.strip(),
    }


def deterministic_verify(xlsx_path: Path, expected_total: int) -> dict[str, Any]:
    required = {"Owner Name", "Property Address", "Total Owed", "Estimated Equity", "Motivation"}
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook["Records"] if "Records" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows)]
    missing = sorted(required - set(headers))
    data = list(rows)
    seen: set[str] = set()
    duplicates = 0
    parcel_idx = headers.index("Parcel REID") if "Parcel REID" in headers else -1
    acct_idx = headers.index("Account ID") if "Account ID" in headers else -1
    for row in data:
        if parcel_idx >= 0 and acct_idx >= 0:
            key = f"{row[parcel_idx]}|{row[acct_idx]}"
            if key in seen:
                duplicates += 1
            seen.add(key)
    failures = []
    if missing:
        failures.append(f"missing columns: {', '.join(missing)}")
    if len(data) != expected_total:
        failures.append(f"row count {len(data)} != expected {expected_total}")
    if duplicates:
        failures.append(f"{duplicates} duplicate parcel/account rows")
    return {"ok": not failures, "failures": failures, "row_count": len(data), "missing_columns": missing, "duplicate_count": duplicates}


def main() -> int:
    parser = argparse.ArgumentParser(description="Manual-only LeadCurate delivery pipeline runner for n8n.")
    parser.add_argument("--market", required=True, choices=sorted(build_delivery.MARKET_REGISTRY))
    parser.add_argument("--lane", required=True)
    parser.add_argument("--count", type=int, default=500)
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    parser.add_argument("--allow-scrape", action="store_true")
    args = parser.parse_args()

    result: dict[str, Any] = {
        "ok": False,
        "market": args.market,
        "lane": args.lane,
        "count": args.count,
        "steps": [],
        "payment_step": {"status": "manual_placeholder", "message": "Payment approval is intentionally manual/no-op."},
    }

    try:
        src = build_delivery.latest_file(args.market, args.lane)
        result["steps"].append({"step": "check_raw_data", "ok": True, "source": str(src)})
    except Exception as exc:
        result["steps"].append({"step": "check_raw_data", "ok": False, "error": str(exc)})
        if not args.allow_scrape:
            result["decision_needed"] = "No raw source found and --allow-scrape was not set. Route to LLM/manual review."
            print(json.dumps(result, indent=2))
            return 2
        scrape = run([sys.executable, str(SCRIPT_DIR / "scrape_dispatcher.py"), "--market", args.market, "--lane", args.lane])
        result["steps"].append({"step": "scrape_dispatcher", **scrape})
        if scrape["returncode"] != 0:
            result["decision_needed"] = "Scrape failed or source not registered. Route to LLM/manual review."
            print(json.dumps(result, indent=2))
            return scrape["returncode"] or 1

    output_dir = Path(args.output_root) / args.market / args.lane
    build = run([sys.executable, str(SCRIPT_DIR / "build_delivery.py"), "--market", args.market, "--lane", args.lane, "--count", str(args.count), "--output-dir", str(output_dir)])
    result["steps"].append({"step": "build_delivery", **build})
    if build["returncode"] != 0:
        result["decision_needed"] = "Build failed or source lacks required signal columns. Route to LLM/manual review."
        print(json.dumps(result, indent=2))
        return build["returncode"] or 1

    payload = json.loads(build["stdout"])
    expected_total = int(payload["total"])
    verify = deterministic_verify(Path(payload["xlsx"]), expected_total)
    result["steps"].append({"step": "verify_delivery", **verify})
    result["delivery"] = payload
    result["ready_to_send"] = bool(verify["ok"])
    result["ok"] = bool(verify["ok"])
    if not verify["ok"]:
        result["decision_needed"] = "Deterministic verify failed. Route to LLM/manual review."
        print(json.dumps(result, indent=2))
        return 3
    print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
