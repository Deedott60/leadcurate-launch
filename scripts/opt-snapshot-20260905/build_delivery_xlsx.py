#!/usr/bin/env python3
"""
Build a branded LeadCurate XLSX delivery file from a package folder.
Cover sheet + one sheet per lane. Customer opens in Excel and works the data.
"""
import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
from openpyxl.formatting.rule import ColorScaleRule

PACKAGES = Path("/opt/leadcurate/packages")

# LeadCurate brand palette
EMERALD = "FF15803D"
EMERALD_SOFT = "FFDFF4E8"
CREAM = "FFFAF7F2"
DARK = "FF0F172A"
STONE = "FFF5F1EB"
INK = "FF1E293B"
MUTED = "FF64748B"
GOLD = "FFD97706"

BORDER_THIN = Border(
    left=Side(style="thin", color="FFE2E8F0"),
    right=Side(style="thin", color="FFE2E8F0"),
    top=Side(style="thin", color="FFE2E8F0"),
    bottom=Side(style="thin", color="FFE2E8F0"),
)

LANE_DESCRIPTIONS = {
    "pre_foreclosure": "Active court-filed foreclosure cases. Closer to sale date = more urgent.",
    "code_violations_open": "Open code violations. Vacant Lot / Vacant Structure / Abandoned flags are highest-priority.",
    "lien_holder_final_orders": "Lien-holder final orders with citation amounts and out-of-state owner flags.",
    "city_lien_active": "Active city liens on Charlotte property. Includes institutional REI fund owners.",
    "vacant_land": "Vacant land parcels with absentee owners, ranked by acreage × land value.",
    "absentee_high_value": "Single-family rentals owned by out-of-state entities, $200k+ assessed.",
}

LANE_DISPLAY_COLS = {
    "pre_foreclosure": ["rank", "score", "property_address", "neighborhood", "sale_date",
                        "days_to_sale", "action_filed_date", "case_number", "parcel_id"],
    "code_violations_open": ["rank", "score", "property_address", "violation_code", "occupancy_status",
                             "inspection_date", "citation_amount", "council_district", "parcel_id"],
    "lien_holder_final_orders": ["rank", "score", "owner_name", "property_address", "owner_mail_city",
                                 "owner_mail_state", "is_out_of_state", "final_order_state",
                                 "final_citation_amount", "date_of_notification", "hearing_scheduled"],
    "city_lien_active": ["rank", "score", "owner_name", "property_address",
                         "property_total_value", "year_built", "heated_sqft",
                         "property_use", "owner_mail_city", "owner_mail_state",
                         "is_out_of_state", "lien_status", "invoice_date",
                         "parcel_id", "parcel_record_url"],
    "vacant_land": ["rank", "score", "owner_name", "property_address", "municipality",
                    "mail_city", "mail_state", "is_absentee_owner", "total_acreage",
                    "land_value", "total_value", "parcel_pid"],
    "absentee_high_value": ["rank", "score", "owner_name", "mailing_address", "mail_city",
                            "mail_state", "mail_zip", "property_location", "property_use",
                            "year_built", "heated_sqft", "land_value", "building_value",
                            "total_value", "is_out_of_state", "parcel_pid"],
}

COL_LABELS = {
    "rank": "Rank", "score": "Score",
    "property_total_value": "Property value",
    "property_building_value": "Building value",
    "property_land_value": "Land value",
    "heated_sqft": "Heated sqft",
    "property_use": "Property use",
    "parcel_record_url": "Record URL",
    "property_address": "Property address", "owner_name": "Owner name",
    "neighborhood": "Neighborhood", "sale_date": "Sale date",
    "days_to_sale": "Days to sale", "action_filed_date": "Action filed",
    "case_number": "Case #", "parcel_id": "Parcel ID",
    "violation_code": "Violation", "occupancy_status": "Occupancy",
    "inspection_date": "Inspected", "citation_amount": "Citation",
    "final_citation_amount": "Citation", "council_district": "Council district",
    "owner_mail_city": "Mail city", "owner_mail_state": "Mail state",
    "is_out_of_state": "Out of state", "final_order_state": "Order state",
    "date_of_notification": "Notified", "hearing_scheduled": "Hearing scheduled",
    "lien_no": "Lien #", "lien_status": "Status",
    "invoice_no": "Invoice #", "invoice_date": "Invoiced",
    "mailing_address": "Owner mailing address",
    "municipality": "Municipality", "mail_city": "Mail city",
    "mail_state": "Mail state", "is_absentee_owner": "Absentee",
    "total_acreage": "Acres", "land_value": "Land value",
    "total_value": "Total value", "parcel_pid": "Parcel ID",
    "mail_zip": "Mail ZIP", "property_location": "Property location",
    "property_use": "Property use", "year_built": "Year built",
    "heated_sqft": "Heated sqft", "building_value": "Building value",
}

CURRENCY_COLS = {"citation_amount", "final_citation_amount", "land_value",
                 "total_value", "building_value", "property_total_value",
                 "property_building_value", "property_land_value"}
INTEGER_COLS = {"rank", "year_built", "heated_sqft", "days_to_sale", "council_district"}
NUMBER_COLS = {"score", "total_acreage"}


def style_cover_sheet(ws, county, state, date, pkg_id, lanes_info):
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 80

    # Brand banner
    ws["B2"] = "LeadCurate."
    ws["B2"].font = Font(name="Calibri", size=36, bold=True, color=DARK)
    ws.merge_cells("B2:C2")
    ws.row_dimensions[2].height = 48

    ws["B3"] = "Better data. Cleaner workflow. No hype."
    ws["B3"].font = Font(name="Calibri", size=11, italic=True, color=MUTED)
    ws.merge_cells("B3:C3")

    # Banner separator
    ws["B5"] = " "
    for col in ("B", "C"):
        ws[f"{col}5"].fill = PatternFill("solid", fgColor=EMERALD)
    ws.row_dimensions[5].height = 4

    # Document title
    ws["B7"] = "COUNTY SEAT DELIVERY"
    ws["B7"].font = Font(name="Calibri", size=10, bold=True, color=EMERALD)
    ws["B7"].alignment = Alignment(horizontal="left")

    ws["B8"] = f"{county}, {state}"
    ws["B8"].font = Font(name="Calibri", size=28, bold=True, color=DARK)
    ws.merge_cells("B8:C8")
    ws.row_dimensions[8].height = 36

    # Meta block
    meta_rows = [
        ("Delivery date", date),
        ("Package ID", pkg_id),
        ("Lane count", str(len(lanes_info))),
        ("Total records", str(sum(m["delivered_rows"] for m in lanes_info))),
    ]
    row = 11
    for label, val in meta_rows:
        ws.cell(row=row, column=2, value=label).font = Font(bold=True, color=MUTED, size=10)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=3, value=val).font = Font(size=12, color=DARK, bold=True)
        row += 1

    # Lanes overview header
    row += 2
    ws.cell(row=row, column=2, value="YOUR LANES THIS MONTH").font = Font(bold=True, color=EMERALD, size=10)
    row += 1
    ws.cell(row=row, column=2, value="Each lane is in its own sheet of this workbook. Click a tab below to work it.").font = Font(italic=True, color=MUTED, size=10)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 2

    headers = ["LANE", "PRODUCT", "RECORDS", "UNIVERSE"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=2 + i, value=h)
        c.font = Font(bold=True, color="FFFFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
    # Adjust columns for this section
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    row += 1
    for i, m in enumerate(lanes_info, 1):
        ws.cell(row=row, column=2, value=f"Lane {i}").font = Font(bold=True, color=EMERALD)
        ws.cell(row=row, column=3, value=m.get("product_name", "Lane"))
        ws.cell(row=row, column=4, value=m.get("delivered_rows", 0))
        ws.cell(row=row, column=5, value=f"{m.get('filtered_universe', m.get('source_total_rows', 0)):,} qualified")
        for col in range(2, 6):
            ws.cell(row=row, column=col).font = Font(color=INK) if col != 2 else Font(bold=True, color=EMERALD)
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=STONE if row % 2 == 0 else "FFFFFFFF")
        row += 1

    # Freshness pitch
    row += 2
    ws.cell(row=row, column=2, value="FRESHNESS").font = Font(bold=True, color=EMERALD, size=10)
    row += 1
    pitch_cell = ws.cell(row=row, column=2,
        value="A name that hits the public record on the first of the month is in your batch within days. The same name does not appear in PropStream's data for another 30–90 days.")
    pitch_cell.font = Font(italic=True, color=DARK, size=12)
    pitch_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=3)
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 20
    ws.row_dimensions[row + 2].height = 20
    row += 4

    # Workflow
    ws.cell(row=row, column=2, value="HOW TO WORK THIS BATCH").font = Font(bold=True, color=EMERALD, size=10)
    row += 1
    workflow = [
        "1. Each lane is its own sheet — click the tabs at the bottom to switch.",
        "2. Records are pre-scored. Rank 1 is highest priority. Work top-down.",
        "3. Mailing state ≠ NC/KY = absentee owner. Higher motivation in most lanes.",
        "4. No phone numbers included — skip-trace through your existing tool (PropStream, BatchLeads, Skip Genie). You stay compliant.",
        "5. Flag exclusions on your side. We'll keep duplicates out of next month's batch.",
    ]
    for line in workflow:
        c = ws.cell(row=row, column=2, value=line)
        c.font = Font(size=10, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 18
        row += 1

    # Compliance
    row += 1
    ws.cell(row=row, column=2, value="COMPLIANCE").font = Font(bold=True, color=EMERALD, size=10)
    row += 1
    comp = ("Property-record data only. Buyer is responsible for owner contact lookup, skip trace, "
            "DNC compliance, TCPA, and outreach decisions. LeadCurate provides data and educational "
            "tools only and does not guarantee deals. Source URLs and pull dates are listed on each lane sheet.")
    c = ws.cell(row=row, column=2, value=comp)
    c.font = Font(size=9, color=MUTED, italic=True)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=3)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 18
    ws.row_dimensions[row + 2].height = 18
    row += 4

    # Footer
    ws.cell(row=row, column=2, value="Your execution closes the deal.").font = Font(italic=True, color=DARK, size=11)
    row += 1
    ws.cell(row=row, column=2, value="— LeadCurate.").font = Font(bold=True, color=EMERALD, size=11)


def style_lane_sheet(ws, lane_meta, rows, lane_key, county_label):
    ws.sheet_properties.tabColor = EMERALD[2:]

    # Top brand bar
    ws["A1"] = "LeadCurate."
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=DARK)
    ws["B1"] = lane_meta.get("product_name", lane_key)
    ws["B1"].font = Font(name="Calibri", size=14, bold=True, color=EMERALD)
    ws.merge_cells("B1:H1")
    ws.row_dimensions[1].height = 28

    desc = LANE_DESCRIPTIONS.get(lane_key, "")
    ws["A2"] = desc
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=MUTED)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 18

    # Meta line
    src_url = lane_meta.get("source_url", "")[:120]
    src_date = lane_meta.get("source_pulled_at", "")
    universe = lane_meta.get("filtered_universe", lane_meta.get("source_total_rows", 0))
    ws["A3"] = (f"County: {county_label}    ·    Records: {len(rows)} of {universe:,} qualified    ·    "
                f"Source pulled: {src_date}    ·    Source: {src_url}")
    ws["A3"].font = Font(name="Calibri", size=9, color=MUTED)
    ws.merge_cells("A3:H3")
    ws.row_dimensions[3].height = 18

    # Separator row
    for col in range(1, 16):
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor=EMERALD)
    ws.row_dimensions[4].height = 4

    # Header row
    cols = LANE_DISPLAY_COLS.get(lane_key, list(rows[0].keys()) if rows else [])
    HEADER_ROW = 6
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=HEADER_ROW, column=i, value=COL_LABELS.get(col, col))
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER_THIN
    ws.row_dimensions[HEADER_ROW].height = 30

    # Data rows
    for ri, r in enumerate(rows, start=HEADER_ROW + 1):
        for ci, col in enumerate(cols, 1):
            raw = r.get(col, "")
            c = ws.cell(row=ri, column=ci)
            if col in CURRENCY_COLS:
                try:
                    c.value = float(str(raw).replace(",", "").replace("$", "")) if raw else None
                    c.number_format = '"$"#,##0'
                except Exception:
                    c.value = raw
            elif col in INTEGER_COLS:
                try:
                    c.value = int(raw) if raw else None
                except Exception:
                    c.value = raw
            elif col in NUMBER_COLS:
                try:
                    c.value = float(raw) if raw else None
                    c.number_format = '0.00'
                except Exception:
                    c.value = raw
            else:
                c.value = raw
            c.font = Font(name="Calibri", size=10, color=INK)
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border = BORDER_THIN
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor=STONE)
            # Rank column highlight
            if col == "rank":
                c.font = Font(name="Calibri", size=10, bold=True, color=EMERALD)
                c.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-width columns based on header + sample data
    for i, col in enumerate(cols, 1):
        letter = get_column_letter(i)
        max_len = max(
            [len(str(COL_LABELS.get(col, col)))] +
            [len(str(r.get(col, ""))) for r in rows[:30]] +
            [10]
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)

    # Freeze header rows
    ws.freeze_panes = f"A{HEADER_ROW + 1}"


def load_lane(pkg_dir, slug):
    lane_dir = pkg_dir / "lanes" / slug
    csv_path = next((p for p in lane_dir.glob("*.csv") if "preview" not in p.name), None)
    meta_path = next(lane_dir.glob("*-meta.json"), None)
    meta = json.loads(meta_path.read_text()) if meta_path else {}
    rows = []
    if csv_path:
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
    return meta, rows


def build_xlsx(pkg_dir, out_path):
    manifest = json.loads((pkg_dir / "manifest.json").read_text())
    county = manifest["customer_county"]
    state = manifest["state"]
    date = manifest["delivery_date"]
    pkg_id = manifest["package_id"]
    lanes_meta = manifest["lanes"]

    wb = Workbook()
    style_cover_sheet(wb.active, county, state, date, pkg_id, lanes_meta)

    SLUG_MAP = {
        "pre_foreclosure": "pre-foreclosure",
        "code_violations_open": "code-violations",
        "lien_holder_final_orders": "lien-holder-orders",
        "city_lien_active": "open-city-liens",
        "vacant_land": "vacant-land-specialty",
        "absentee_high_value": "high-value-absentee",
    }
    LANE_TAB_NAMES = {
        "pre_foreclosure": "Pre-Foreclosure",
        "code_violations_open": "Code Violations",
        "lien_holder_final_orders": "Lien Orders",
        "city_lien_active": "City Liens",
        "vacant_land": "Vacant Land",
        "absentee_high_value": "Absentee SF",
    }

    for lm in lanes_meta:
        lane_key = lm["lane"]
        slug = SLUG_MAP.get(lane_key, lane_key.replace("_", "-"))
        meta, rows = load_lane(pkg_dir, slug)
        tab = LANE_TAB_NAMES.get(lane_key, lane_key)
        ws = wb.create_sheet(title=tab)
        style_lane_sheet(ws, meta if meta else lm, rows, lane_key, f"{county}, {state}")

    wb.save(out_path)
    print(f"  -> {out_path}  ({out_path.stat().st_size} bytes)")


if __name__ == "__main__":
    targets = sys.argv[1:] or [
        ("louisville-ky-2026-06-19", "Louisville-KY-County-Seat-Delivery-2026-06-19.xlsx"),
        ("charlotte-nc-2026-06-19", "Charlotte-NC-County-Seat-Delivery-2026-06-19.xlsx"),
    ]
    for pid, fname in targets:
        pdir = PACKAGES / pid
        if not pdir.exists():
            print(f"!! not found: {pdir}")
            continue
        print(f"\n=== {pid} ===")
        out = pdir / fname
        build_xlsx(pdir, out)
