#!/usr/bin/env python3
"""Build branded Houston XLSX on the VPS (openpyxl available)."""
import csv
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

SNAP = Path("/opt/leadcurate/snapshots/harris-tx/2026-06-21")
csv_path = SNAP / "harris-tx-permit-burnout-2026-06-21.csv"
meta = json.loads((SNAP / "meta.json").read_text())

EMERALD = "FF15803D"
EMERALD_DARK = "FF14532D"
CREAM = "FFFAF7F2"
CREAM_2 = "FFF3EDDF"
DARK = "FF0F172A"
SLATE = "FF475569"
SLATE_LIGHT = "FF94A3B8"
GOLD = "FFB45309"
CRIMSON = "FF991B1B"

wb = Workbook()

# === Sheet 1: COVER ===
cover = wb.active
cover.title = "LeadCurate"
cover.sheet_view.showGridLines = False
cover.column_dimensions["A"].width = 4
cover.column_dimensions["B"].width = 36
cover.column_dimensions["C"].width = 36
cover.column_dimensions["D"].width = 18
cover.column_dimensions["E"].width = 4

# Brand block
cover["B2"] = "L"
cover["B2"].font = Font(name="Calibri", size=28, bold=True, color="FFFAF7F2")
cover["B2"].fill = PatternFill("solid", fgColor=EMERALD)
cover["B2"].alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[2].height = 40
cover["C2"] = "LeadCurate"
cover["C2"].font = Font(name="Calibri", size=22, bold=True, color=DARK)
cover["C2"].alignment = Alignment(vertical="center")

cover["B5"] = "DISCOVERY SNAPSHOT · HOUSTON METRO"
cover["B5"].font = Font(name="Calibri", size=10, bold=True, color=EMERALD)
cover["B6"] = "Harris County, TX"
cover["B6"].font = Font(name="Calibri", size=28, bold=True, color=DARK)
cover["B7"] = "Permit Burnout Lane"
cover["B7"].font = Font(name="Calibri", size=22, bold=True, color=DARK)
cover.row_dimensions[6].height = 34
cover.row_dimensions[7].height = 28

cover["B9"] = (
    f"Single-family residential parcels carrying active or recent distress permits — "
    f"FIRELOSS, DEMOLITION, EMERGENCY REPAIR, STORM, FLOOD. Cross-referenced with "
    f"absentee mailing addresses and entity ownership. Top {meta['delivered_top_n']:,} "
    f"ranked by distress score (0–110)."
)
cover["B9"].font = Font(name="Calibri", size=11, color=SLATE)
cover["B9"].alignment = Alignment(wrap_text=True, vertical="top")
cover.merge_cells("B9:D9")
cover.row_dimensions[9].height = 72

cover["B11"] = f"Source: HCAD bulk extract ({meta['pulled']}) · Processed: {meta['processed']}"
cover["B11"].font = Font(name="Consolas", size=9, color=SLATE_LIGHT)
cover.merge_cells("B11:D11")

# Stats table
stats = [
    ("Total HCAD universe", f"{meta['universe_size']:,}", "parcels"),
    ("Candidates w/ permit signal", f"{meta['candidates_with_permits']:,}", "active/recent"),
    ("Residential qualified", f"{meta['qualified_residential']:,}", "SFR / dup / townhouse"),
    ("Delivered top-N", f"{meta['delivered_top_n']:,}", "ranked by score"),
    ("Fire-loss permits", f"{meta['fire_loss_count']:,}", "verified FIRELOSS"),
    ("Demolition permits", f"{meta['demolition_count']:,}", "DEMO filed"),
    ("Repair / damage permits", f"{meta['repair_damage_count']:,}", "storm/flood/structural"),
    ("Absentee owners", f"{meta['absentee_count']:,}", "mailing ≠ situs"),
    ("Entity-owned", f"{meta['entity_count']:,}", "LLC/Trust/Corp/REIT"),
    ("Avg market value", f"${meta['avg_market_value']:,}", "HCAD assessed"),
    ("Score range", f"{meta['min_score']}–{meta['max_score']}", "out of 110"),
]
r = 14
cover[f"B{r-1}"] = "THE FUNNEL"
cover[f"B{r-1}"].font = Font(name="Calibri", size=10, bold=True, color=EMERALD)

for label, value, detail in stats:
    cover[f"B{r}"] = label
    cover[f"B{r}"].font = Font(name="Calibri", size=10, color=SLATE)
    cover[f"C{r}"] = value
    cover[f"C{r}"].font = Font(name="Calibri", size=14, bold=True, color=DARK)
    cover[f"D{r}"] = detail
    cover[f"D{r}"].font = Font(name="Calibri", size=9, color=SLATE_LIGHT)
    cover.row_dimensions[r].height = 22
    r += 1

# Footer note
r += 2
cover[f"B{r}"] = "WHAT'S INCLUDED"
cover[f"B{r}"].font = Font(name="Calibri", size=10, bold=True, color=EMERALD)
includes = [
    "Owner name + entity classification (LLC/Trust/Corp)",
    "Full mailing address with out-of-state flag",
    "Site address: street, unit, city, ZIP",
    "Distress signal kinds (FIRELOSS, FLOOD, DEMO, REPAIR…)",
    "Latest permit year + description text",
    "Market value, building value, land value (HCAD)",
    "Year built, building sq ft, land sq ft",
    "Active permit count per parcel",
    "Distress score 0–110 for ranking",
]
for s in includes:
    r += 1
    cover[f"B{r}"] = f"  ✓ {s}"
    cover[f"B{r}"].font = Font(name="Calibri", size=11, color=DARK)
    cover.merge_cells(f"B{r}:D{r}")
    cover.row_dimensions[r].height = 18

# === Sheet 2: DATA ===
ws = wb.create_sheet("Houston Permit Burnout")
with open(csv_path, newline="", encoding="utf-8") as fp:
    reader = csv.reader(fp)
    headers = next(reader)
    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h.replace("_", " ").title())
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=EMERALD)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    thin = Side(style="thin", color="FFE2DCCF")
    border = Border(bottom=thin)
    def coerce(v):
        if v is None or v == "":
            return v
        try:
            if "." in v:
                return float(v)
            return int(v)
        except (ValueError, TypeError):
            return v

    for row_i, row in enumerate(reader, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=row_i, column=c, value=coerce(v))
            cell.font = Font(name="Calibri", size=10, color=DARK)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Money formatting
            if c in (15, 16, 17):  # mkt_val, land_val, bld_val
                cell.number_format = '"$"#,##0'

# Column widths
widths = [14, 28, 9, 9, 28, 8, 12, 8, 14, 7, 8, 7, 9, 10, 12, 12, 12, 7, 22, 7, 28, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Save
out_path = SNAP / "Houston-Permit-Burnout-2026-06-21.xlsx"
wb.save(str(out_path))
print(f"Wrote: {out_path}  ({out_path.stat().st_size:,} bytes)")
