#!/usr/bin/env python3
"""Process Jefferson AL DelinquentParcelList.xls → CSV snapshot."""
import csv
import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "xlrd", "--break-system-packages", "-q"])
    from openpyxl import load_workbook

XLS = Path("/opt/leadcurate/raw_imports/jefferson-al/2026-06-21/DelinquentParcelList.xls")
OUT_DIR = Path("/opt/leadcurate/snapshots/jefferson-al") / date.today().isoformat()
OUT_DIR.mkdir(parents=True, exist_ok=True)

ENTITY_PATTERNS = re.compile(
    r"\b(LLC|L\.L\.C\.|INC|CORP|CORPORATION|CO|COMPANY|LP|LTD|TRUST|TRUSTEE|"
    r"REIT|HOLDINGS?|PARTNERS?|GROUP|ASSOC|ASSN|FOUNDATION|FUND|CHURCH|"
    r"HOA|REALTY|PROPERTIES|INVESTMENTS?|RENTALS?|MGT|MGMT|"
    r"VENTURE|EQUITY|CAPITAL|SFR|REI|HOMES?|ESTATE)\b", re.IGNORECASE)


# First detect file type — .xls extension might actually be HTML, OOXML, or true XLS
data = XLS.read_bytes()
head = data[:8]
print(f"File magic bytes: {head}", file=sys.stderr)

rows = []
header = None

# Try OOXML (zip) - real .xlsx
if head.startswith(b"PK"):
    print("  detected: XLSX (zip)", file=sys.stderr)
    wb = load_workbook(str(XLS), read_only=True, data_only=True)
    ws = wb.active
    for r in ws.iter_rows(values_only=True):
        if not any(c is not None and str(c).strip() for c in r):
            continue
        cells = ["" if c is None else str(c).strip() for c in r]
        if header is None:
            header = cells
        else:
            rows.append(cells)

# Try CFB (true old XLS)
elif head.startswith(b"\xd0\xcf\x11\xe0"):
    print("  detected: legacy XLS (CFB)", file=sys.stderr)
    try:
        import xlrd
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "xlrd==1.2.0", "--break-system-packages", "-q"])
        import xlrd
    wb = xlrd.open_workbook(str(XLS))
    ws = wb.sheet_by_index(0)
    for ri in range(ws.nrows):
        cells = [str(ws.cell_value(ri, ci)).strip() for ci in range(ws.ncols)]
        if header is None:
            header = cells
        else:
            rows.append(cells)

# Maybe it's HTML masquerading
elif b"<html" in head.lower() or b"<table" in data[:200].lower():
    print("  detected: HTML masquerading as .xls", file=sys.stderr)
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "beautifulsoup4", "--break-system-packages", "-q"])
        from bs4 import BeautifulSoup
    soup = BeautifulSoup(data, "html.parser")
    for tbl in soup.find_all("table"):
        for tr in tbl.find_all("tr"):
            cells = [td.get_text(" ", strip=True) for td in tr.find_all(["td", "th"])]
            if not any(cells):
                continue
            if header is None:
                header = cells
            else:
                rows.append(cells)

else:
    print(f"  unknown format. First 200 bytes: {data[:200]}", file=sys.stderr)
    sys.exit(1)

print(f"  header: {header}", file=sys.stderr)
print(f"  data rows: {len(rows)}", file=sys.stderr)

if not rows:
    print("WARNING: 0 rows", file=sys.stderr)
    sys.exit(0)

# Build dict rows
dicts = []
for r in rows:
    if header and len(r) >= len(header):
        dicts.append(dict(zip(header, r[:len(header)])))

# Detect columns
def find_col(header, *patterns):
    for h in header:
        low = h.lower()
        for p in patterns:
            if p in low:
                return h
    return None

owner_col = find_col(header, "owner", "name", "taxpayer")
parcel_col = find_col(header, "parcel", "pin", "id")
addr_col = find_col(header, "address", "situs", "location", "property")
amount_col = find_col(header, "totaldue", "amountdue", "duetotal", "balance", "amount due")
if not amount_col:
    amount_col = find_col(header, "due", "balance", "amount")
if not amount_col:
    amount_col = find_col(header, "total")
year_col = find_col(header, "year")

print(f"  cols: owner={owner_col} parcel={parcel_col} addr={addr_col} amt={amount_col} yr={year_col}", file=sys.stderr)


def parse_amount(v):
    try:
        return float(re.sub(r"[^\d.]", "", str(v or "0")))
    except Exception:
        return 0.0


# Score and emit
out = []
for d in dicts:
    owner = d.get(owner_col, "") if owner_col else ""
    if not owner:
        continue
    parcel = d.get(parcel_col, "") if parcel_col else ""
    addr = d.get(addr_col, "") if addr_col else ""
    amount = parse_amount(d.get(amount_col, "")) if amount_col else 0.0
    year = d.get(year_col, "") if year_col else ""
    is_entity = bool(ENTITY_PATTERNS.search(owner))
    score = 50
    if amount > 10000: score += 25
    elif amount > 5000: score += 18
    elif amount > 2500: score += 12
    elif amount > 1000: score += 6
    if is_entity: score += 8
    out.append({
        "parcel_id": parcel,
        "owner_name": owner,
        "owner_type": "entity" if is_entity else "individual",
        "site_addr": addr,
        "delinquent_amount": amount,
        "tax_year": year,
        "score": score,
        **{k: v for k, v in d.items() if k not in (owner_col, parcel_col, addr_col, amount_col, year_col)},
    })

out.sort(key=lambda r: (r["score"], r["delinquent_amount"]), reverse=True)

csv_path = OUT_DIR / f"jefferson-al-delinquent-{date.today().isoformat()}.csv"
with open(csv_path, "w", newline="", encoding="utf-8") as fp:
    if out:
        writer = csv.DictWriter(fp, fieldnames=list(out[0].keys()))
        writer.writeheader()
        writer.writerows(out)


def redact_name(n):
    parts = n.split()
    return " ".join((p[0] + "***") if len(p) > 1 else p for p in parts)


def redact_addr(a):
    return re.sub(r"\b\d+\b", "###", a, count=1)


preview_path = OUT_DIR / f"jefferson-al-delinquent-{date.today().isoformat()}-preview.csv"
with open(preview_path, "w", newline="", encoding="utf-8") as fp:
    if out:
        writer = csv.DictWriter(fp, fieldnames=list(out[0].keys()))
        writer.writeheader()
        for r in out[:25]:
            rr = dict(r)
            rr["owner_name"] = redact_name(rr["owner_name"])
            rr["site_addr"] = redact_addr(rr["site_addr"])
            writer.writerow(rr)

ent = sum(1 for r in out if r["owner_type"] == "entity")
total_value = sum(r["delinquent_amount"] for r in out)
big_balance = sum(1 for r in out if r["delinquent_amount"] >= 5000)
meta = {
    "market": "Jefferson County, AL (Birmingham metro)",
    "lane": "Tax Delinquent (Birmingham Division)",
    "pulled": date.today().isoformat(),
    "processed": date.today().isoformat(),
    "source": "Jefferson County Capture CAMA portal — DelinquentParcelList.xls",
    "vendor": "Capture CAMA (jeffersonexpress.capturecama.com)",
    "api_endpoint": "POST /SearchDelq",
    "total_rows": len(out),
    "entity_count": ent,
    "individual_count": len(out) - ent,
    "high_balance_5k_plus": big_balance,
    "total_delinquent_value": int(total_value),
    "avg_delinquent_amount": int(total_value / max(1, len(out))),
    "min_score": out[-1]["score"] if out else 0,
    "max_score": out[0]["score"] if out else 0,
    "header_detected": header,
}
(OUT_DIR / "meta.json").write_text(json.dumps(meta, indent=2))

print(f"\n=== DONE ===", file=sys.stderr)
print(f"  CSV:     {csv_path}", file=sys.stderr)
print(f"  Preview: {preview_path}", file=sys.stderr)
print(json.dumps(meta, indent=2))
